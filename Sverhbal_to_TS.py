import config as cfg
import cx_Oracle as ora
import pandas as pd
import os
from StyleFrame import StyleFrame, Styler, utils
from openpyxl.styles import Border, Side, Alignment
from openpyxl import load_workbook
from time import time

import smtplib
from email.mime.application import MIMEApplication
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText

time_start = time()

# Подключение к БД
conn_sib = ora.connect(cfg.user_db + '/' + cfg.pass_db + '@' + cfg.db_sib, encoding='utf-8')
conn_eur = ora.connect(cfg.user_db + '/' + cfg.pass_db + '@' + cfg.db_eur, encoding='utf-8')
# Загрузка даты отчета
ye = cfg.year
mon = cfg.mon
date = '01.' + mon + '.' + ye
print('Отчетный месяц: ', date)
# Список месяцев
m = ['январь', 'февраль', 'март', 'апрель', 'май', 'июнь',
     'июль', 'август', 'сентябрь', 'октябрь', 'ноябрь', 'декабрь']
# Корневая папка
path_0 = cfg.path
path_1 = path_0 + 'Отчеты коллегам/' + ye + '/' + mon
path_DD = path_0 + 'Отчеты коллегам/' + ye + '/' + mon + '/' + 'ДД в НЦЗ.xlsx'
path_sverhbal = path_0 + 'Отчеты коллегам/' + ye + '/' + mon + '/' + 'Объем сверхбаланса ' + m[int(mon) - 1] + ' ' + ye + '.xlsx'
path_fact_DV = path_0 + 'Отчеты коллегам/' + ye + '/' + mon + '/' + 'Суммарный факт по ГТП Дальнего востока ' + m[int(mon) - 1] + ' ' + ye + '.xlsx'

if not os.path.exists(path_1):
    os.makedirs(path_1)

# Данные по объёмам сверхбаланса
SverhBal_eur = '''Select g.target_date as Дата, 'Европа' as ЦЗ, g.gen as Генерация, p.potr as Потребление
                           from
                                (select trunc(x.target_date,'month') target_date,sum(x.dd_max_gen) gen
                                   from
                                    (select distinct g.target_date, g.station_id, g.dd_max_gen
                                       from frsdb_dev.ncz_dd_volume g
                                       where g.target_date between to_date(:d, 'dd.mm.yyyy') and last_day(to_date(:d, 'dd.mm.yyyy'))
                                       and g.end_ver>999999999) x group by trunc(x.target_date,'month')) g,
                                
                                (select trunc(x.target_date,'month')target_date, sum(x.dd_max_con) potr
                                from
                                     (select distinct g.target_date, g.con_gtp_id, g.dd_max_con
                                     from frsdb_dev.ncz_dd_volume g
                                     where g.target_date between to_date(:d, 'dd.mm.yyyy') and last_day(to_date(:d, 'dd.mm.yyyy'))
                                     and g.end_ver>999999999
                                     ) x group by trunc(x.target_date,'month')) p'''

SverhBal_sib = '''Select g.target_date as Дата, 'Сибирь' as ЦЗ, g.gen as Генерация, p.potr as Потребление
                           from
                                (select trunc(x.target_date,'month') target_date,sum(x.dd_max_gen) gen
                                   from
                                    (select distinct g.target_date, g.station_id, g.dd_max_gen
                                       from frsdb_dev_sib.ncz_dd_volume g
                                       where g.target_date between to_date(:d, 'dd.mm.yyyy') and last_day(to_date(:d, 'dd.mm.yyyy'))
                                       and g.end_ver>999999999) x group by trunc(x.target_date,'month')) g,

                                (select trunc(x.target_date,'month')target_date, sum(x.dd_max_con) potr
                                from
                                     (select distinct g.target_date, g.con_gtp_id, g.dd_max_con
                                     from frsdb_dev_sib.ncz_dd_volume g
                                     where g.target_date between to_date(:d, 'dd.mm.yyyy') and last_day(to_date(:d, 'dd.mm.yyyy'))
                                     and g.end_ver>999999999
                                     ) x group by trunc(x.target_date,'month')) p'''

# Суммарный факт по ДВ

Fact_DV = ''' Select p.target_date as Дата, potr as Потребление, gen as Генерация
                From
                    (SELECT distinct trunc(a.target_date, 'month') target_date, sum(a.fact) over (partition by trunc(a.target_date, 'month')) potr
                       FROM frsdb_dev_sib.ncz_con_volume a
                       where a.target_date between to_date(:d, 'dd.mm.yyyy') and last_day(to_date(:d, 'dd.mm.yyyy'))
                       and a.end_ver>9999999999999
                       and a.is_daily = 0) p,

                (SELECT distinct trunc(a.target_date, 'month') target_date, sum(a.fact) over (partition by trunc(a.target_date, 'month')) gen
                   FROM frsdb_dev_sib.ncz_gen_volume a
                    where a.target_date between to_date(:d, 'dd.mm.yyyy') and last_day(to_date(:d, 'dd.mm.yyyy'))
                    and a.end_ver>9999999999999
                    and a.is_daily = 0)g'''


df_SverhBal_eur = pd.read_sql(SverhBal_eur, conn_eur, params={'d': date})
df_SverhBal_sib = pd.read_sql(SverhBal_sib, conn_sib, params={'d': date})
df_Fact_DV = pd.read_sql(Fact_DV, conn_sib, params={'d': date})

conn_eur.close()
conn_sib.close()

# Объединяем массивы данных по сверхбалансу и приводим дату в номальный вид
df_SverhBal = df_SverhBal_eur.append(df_SverhBal_sib)
df_SverhBal['ДАТА'] = df_SverhBal['ДАТА'].astype('str')
df_Fact_DV['ДАТА'] = df_Fact_DV['ДАТА'].astype('str')


# Экспортируем выгруженные данные в xlsx
df_SverhBal.to_excel(path_sverhbal, index=False)
df_Fact_DV.to_excel(path_fact_DV, index=False)

# Создаем шаблоны стилей
border = Border(left=Side(border_style='thin', color='FF000000'),
                right=Side(border_style='thin', color='FF000000'),
                top=Side(border_style='thin', color='FF000000'),
                bottom=Side(border_style='thin', color='FF000000'))
align_head = Alignment(horizontal='center', vertical='center',
                       text_rotation=0, wrap_text=True,
                       shrink_to_fit=True, indent=0)
align_cell = Alignment(horizontal='right', vertical='center',
                       text_rotation=0, wrap_text=False,
                       shrink_to_fit=False, indent=0)

# Форматирование Excel
def exstyle(path):
    # Открываем рабочую страницу основного отчета
    wb = load_workbook(path)
    ws = wb.active
    # Устанавливаем ширину столбцов
    dim = {}
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            if cell.value:
                dim[cell.column] = max((dim.get(cell.column, 0), len(str(cell.value))))
    for col, value in dim.items():
        ws.column_dimensions[col].width = 8 + 0.85 * value
    # Задаем стили ячеек
    for row in ws.iter_rows():
        for cell in row:
            if cell in ws['1:1']:
                cell.border = border
                cell.alignment = align_head
            else:
                cell.border = border
                cell.alignment = align_cell
    wb.save(path)

exstyle(path_sverhbal)
exstyle(path_fact_DV)

# Формирование письма
send_body = 'Добрый день!\n\nОтправляем фактические объемы ДД, сверхбалансовые величины и факт по ГТП Дальнего Востока ' \
            'за ' + m[int(mon) - 1] + ' ' + ye + '.\n\n' \
            'С уважением, \nОтдел расчета объемов покупки и \nпродажи электрической энергии АО «АТС»'
msg = MIMEMultipart()
msg['From'] = cfg.send_from
msg['To'] = cfg.send_to_SverhBal
msg['CC'] = cfg.send_cc
msg['Subject'] = 'Сверхбалансовые величины и факт по ГТП за ' + m[int(mon) - 1] + ' ' + ye
msg.attach(MIMEText(send_body, 'plain'))


def add_file(path):
    with open(path, "rb") as f:
        part = MIMEApplication(f.read())
    part.add_header('Content-Disposition', 'attachment', filename=os.path.basename(path))
    msg.attach(part)


add_file(path_sverhbal)
add_file(path_fact_DV)
add_file(path_DD)
# Отправка письма
s = smtplib.SMTP('smtp.rosenergo.com')
s.send_message(msg)
s.quit()

print('Письмо отправлено за: ', round(time() - time_start, 2), 'сек')