import zipfile
import config as cfg
import os
import shutil
import cx_Oracle as ora
import pandas as pd
from openpyxl.styles import Border, Side, Alignment
from openpyxl import load_workbook

from time import time

time_start = time()

# Подключение к БД
conn_eur = ora.connect(cfg.user_db + '/' + cfg.pass_db + '@' + cfg.db_eur)
conn_sib = ora.connect(cfg.user_db + '/' + cfg.pass_db + '@' + cfg.db_sib)

# Загрузка даты отчета
ye = cfg.year
mon = cfg.mon
date = '01.' + mon + '.' + ye
print('Отчетный месяц: ', date)

# Список месяцев
m = ['январь', 'февраль', 'март', 'апрель', 'май', 'июнь',
     'июль', 'август', 'сентябрь', 'октябрь', 'ноябрь', 'декабрь']
# Корневая папка
path_0 = cfg.path
# Собираем пути к выгруженным отчётам
path_eur = path_0 + 'Отчеты для фин гарантий/' + ye + '/' + mon + '/' + ye + mon + '01_report_vc_ppp_eur' + '.xml'
path_sib = path_0 + 'Отчеты для фин гарантий/' + ye + '/' + mon + '/' + ye + mon + '01_report_vc_ppp_sib' + '.xml'

# Обратный путь для перекладывания архивов
path_out_eur = path_0 + 'Отчеты для фин гарантий/' + ye + '/' + mon + '/' + ye + mon + '01_report_vc_ppp_eur' + '.zip'
path_out_sib = path_0 + 'Отчеты для фин гарантий/' + ye + '/' + mon + '/' + ye + mon + '01_report_vc_ppp_sib' + '.zip'

# Убираем толстые файлы в архивы
PppEur = zipfile.ZipFile(ye + mon + '01_report_vc_ppp_eur.zip', 'w', zipfile.ZIP_DEFLATED)
eur = os.path.abspath(path_eur)
PppEur.write(eur, ye + mon + '01_report_vc_ppp_eur' + '.xml')
PppEur.close()

PppSib = zipfile.ZipFile(ye + mon + '01_report_vc_ppp_sib.zip', 'w', zipfile.ZIP_DEFLATED)
sib = os.path.abspath(path_sib)
PppSib.write(sib, ye + mon + '01_report_vc_ppp_sib' + '.xml')
PppSib.close()

# Перекладываем архивы в нужную папку
shutil.move(ye + mon + '01_report_vc_ppp_eur.zip', path_out_eur)
shutil.move(ye + mon + '01_report_vc_ppp_sib.zip', path_out_sib)
# Вытаскиваем данные по РД для особых регионов

rd_zap_eur = '''select distinct p.gtp_id, p.gtp_code, p.target_date, p.hour, p.volume_rd
               from frsdb_dev.dev_plan p, FRSDB_DEV.trader_xattr t
              where p.end_ver>999999999999 
                    and p.target_Date between to_date(:d, 'dd.mm.yyyy') and last_day(to_date(:d, 'dd.mm.yyyy'))
                    and p.dir=1 
                    and t.real_trader_id=p.gtp_id 
                    and t.xattr_type='rd_specific_consumer'
                    and t.xattr_value=1
                    and p.target_date between t.begin_date and t.end_date
                  order by 3,4,1'''
rd_eur = pd.read_sql(rd_zap_eur, conn_eur, params={'d': date})

rd_zap_sib = '''select distinct p.gtp_id, p.gtp_code, p.target_date, p.hour, p.volume_rd
               from frsdb_dev_sib.dev_plan p, FRSDB_DEV_sib.trader_xattr t
              where p.end_ver>999999999999 
                    and p.target_Date between to_date(:d, 'dd.mm.yyyy') and last_day(to_date(:d, 'dd.mm.yyyy'))
                    and p.dir=1 
                    and t.real_trader_id=p.gtp_id 
                    and t.xattr_type='rd_specific_consumer'
                    and t.xattr_value=1
                    and p.target_date between t.begin_date and t.end_date
                  order by 3,4,1'''
rd_sib = pd.read_sql(rd_zap_sib, conn_sib, params={'d': date})
conn_eur.close()
conn_sib.close()
# Объединяем данные Европы и Сибири в один DataFrame, затем добавляем столбец с ценовой зоной,
# меняем порядок столбцов, чтобы ЦЗ оказалась в начале и убираем из даты время
rd_eur['Ценовая зона'] = 1
rd_sib['Ценовая зона'] = 2
rd = rd_eur.append(rd_sib)
# cols = rd.columns.tolist()
cols = ['Ценовая зона', 'GTP_ID', 'GTP_CODE', 'TARGET_DATE', 'HOUR', 'VOLUME_RD']
rd = rd[cols]
rd['TARGET_DATE'] = rd['TARGET_DATE'].dt.date

# Экспортируем выгруженные данные в xlsx
path_rd = path_0 + 'Отчеты для фин гарантий/' + 'Для проверки РД по особым/' + ye + '/' + 'РД_' + mon + ye + '.xlsx'
rd.to_excel(path_rd, index=False)

# Создаем шаблоны стилей
border = Border(left=Side(border_style='thin', color='FF000000'),
                right=Side(border_style='thin', color='FF000000'),
                top=Side(border_style='thin', color='FF000000'),
                bottom=Side(border_style='thin', color='FF000000'))
align_head = Alignment(horizontal='center', vertical='center',
                       text_rotation=0, wrap_text=True,
                       shrink_to_fit=True, indent=0)
align_cell = Alignment(horizontal='right', vertical='center',
                       text_rotation=0, wrap_text=False,
                       shrink_to_fit=False, indent=0)


# Форматирование Excel
def exstyle(path):
    # Открываем рабочую страницу основного отчета
    wb = load_workbook(path)
    ws = wb.active
    # Устанавливаем ширину столбцов
    dim = {}
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            if cell.value:
                dim[cell.column] = max((dim.get(cell.column, 0), len(str(cell.value))))
    for col, value in dim.items():
        ws.column_dimensions[col].width = 8 + 0.85 * value
    # Задаем стили ячеек
    # for row in ws.iter_rows():
    #     for cell in row:
    #         if cell in ws['1:1']:
    #             cell.border = border
    #             cell.alignment = align_head
    #         else:
    #             cell.border = border
    #             cell.alignment = align_cell
    wb.save(path)


exstyle(path_rd)

# Конечные пути файлов для отпраки:
# Для архивов смотри в Перекладывании выше
# Для экселя по особым в создании xlsx файла описан путь
# Добавляем путь для xml по НЦЗ, тк он только выгружается со стенда и не подвергается никакой обработке
path_ncz = path_0 + 'Отчеты для фин гарантий/' + ye + '/' + mon + '/' + ye + mon + '01_report_ncz_vc_ppp' + '.xml'

print(round(time() - time_start, 2), 'sec')
