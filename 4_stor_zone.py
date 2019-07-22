import config as cfg
import cx_Oracle as ora
import pandas as pd
import os
from openpyxl.styles import Border, Side, Alignment
from openpyxl import load_workbook
from time import time

import smtplib
from email.mime.application import MIMEApplication
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText

time_start = time()

# Подключение к БД
conn_sib = ora.connect(cfg.user_db + '/' + cfg.pass_db + '@' + cfg.db_sib, encoding='utf-8')
conn_eur = ora.connect(cfg.user_db + '/' + cfg.pass_db + '@' + cfg.db_eur, encoding='utf-8')
# Загрузка даты отчета
ye = cfg.year
mon = cfg.mon
date = '01.' + mon + '.' + ye
print('Отчетный месяц: ', date)
# Список месяцев
m = ['январь', 'февраль', 'март', 'апрель', 'май', 'июнь',
     'июль', 'август', 'сентябрь', 'октябрь', 'ноябрь', 'декабрь']
# Корневая папка
path_0 = cfg.path
path_1 = path_0 + 'Отчеты коллегам/' + ye + '/' + mon
path_FourSt_Zone = path_1 +  '/' + '4-х сторонние договоры по зонам ' + m[int(mon) - 1] + ' ' + ye + '.xlsx'

if not os.path.exists(path_1):
    os.makedirs(path_1)

# Данные по 4-х стороннам договорам
Four_stor_eur = '''SELECT CASE con.zone
            WHEN 1 THEN 'АРХАНГЕЛЬСК' 
            WHEN 2 THEN 'КАЛИНИНГРАД'
            WHEN 3 THEN 'КОМИ' END ЗОНА,
            nvl(con,0) + nvl(gen,0) + nvl(fsk,0)+nvl(imp_gen,0)+nvl(imp_con,0) AS ОБЪЕМ, nvl(gen,0) AS ГЕНЕРАЦИЯ, nvl(con,0) AS 
ПОТРЕБЛЕНИЕ, nvl(fsk,0) AS ФСК, nvl(imp_gen,0) as ИМПОРТ , nvl(imp_con,0) as ЭКСПОРТ
  FROM (SELECT   TO_CHAR (a.target_date, 'MONTH') AS mon,
                 SUM (a.buy_volume) AS con, tr.is_unpriced_zone AS 
ZONE
            FROM frsdb_dev.ncz_plan_con_volume a, 
frsdb_dev.trader tr
           WHERE a.target_date BETWEEN tr.begin_date AND 
tr.end_date
             AND a.end_ver > 9999999
             AND a.target_date BETWEEN to_date(:d, 'dd.mm.yyyy') and last_day(to_date(:d, 'dd.mm.yyyy'))
             AND a.gtp_id = tr.real_trader_id
        GROUP BY TO_CHAR (a.target_date, 'MONTH'), 
tr.is_unpriced_zone) con,
       (SELECT   TO_CHAR (a.target_date, 'MONTH') AS mon,
                 SUM (a.sell_volume) AS gen, tr.is_unpriced_zone AS 
ZONE
            FROM frsdb_dev.ncz_plan_gen_volume a, 
frsdb_dev.trader tr
           WHERE a.target_date BETWEEN tr.begin_date AND 
tr.end_date
             AND a.end_ver > 9999999
             AND a.target_date BETWEEN to_date(:d, 'dd.mm.yyyy') and last_day(to_date(:d, 'dd.mm.yyyy'))
             AND a.gtp_id = tr.real_trader_id
        GROUP BY TO_CHAR (a.target_date, 'MONTH'), 
tr.is_unpriced_zone) gen,
       (SELECT   TO_CHAR (a.target_date, 'MONTH') AS mon, SUM 
(a.buy_volume)
                                                                       AS fsk,
                 tr.is_unpriced_zone AS ZONE
            FROM frsdb_dev.ncz_plan_fsk_volume a, 
frsdb_dev.trader tr
           WHERE a.target_date BETWEEN tr.begin_date AND 
tr.end_date
             AND a.end_ver > 9999999
             AND a.target_date BETWEEN to_date(:d, 'dd.mm.yyyy') and last_day(to_date(:d, 'dd.mm.yyyy'))
            and a.fsk_id=tr.real_trader_id
         GROUP BY TO_CHAR (a.target_date, 'MONTH'), 
tr.is_unpriced_zone) fsk,
          (SELECT   TO_CHAR (a.target_date, 'MONTH') AS mon, 
SUM (a.impex_volume)
                                                                       AS imp_gen,
                 tr.is_unpriced_zone AS ZONE
            FROM frsdb_dev.ncz_plan_impex_volume a, 
frsdb_dev.trader tr
           WHERE a.target_date BETWEEN tr.begin_date AND 
tr.end_date
             AND a.end_ver > 9999999
             AND a.target_date BETWEEN to_date(:d, 'dd.mm.yyyy') and last_day(to_date(:d, 'dd.mm.yyyy'))
             AND a.gtp_id = tr.real_trader_id
             and a.dir=2
         GROUP BY  TO_CHAR (a.target_date, 'MONTH'), 
tr.is_unpriced_zone) imp_gen,
          (SELECT   TO_CHAR (a.target_date, 'MONTH') AS mon, 
SUM (a.impex_volume)
                                                                       AS imp_con,
                 tr.is_unpriced_zone AS ZONE
            FROM frsdb_dev.ncz_plan_impex_volume a, 
frsdb_dev.trader tr
           WHERE a.target_date BETWEEN tr.begin_date AND 
tr.end_date
             AND a.end_ver > 9999999
             AND a.target_date BETWEEN to_date(:d, 'dd.mm.yyyy') and last_day(to_date(:d, 'dd.mm.yyyy'))
             AND a.gtp_id = tr.real_trader_id
             and a.dir=1
         GROUP BY  TO_CHAR (a.target_date, 'MONTH'), 
tr.is_unpriced_zone) imp_con
 WHERE con.mon = gen.mon
   AND fsk.mon = gen.mon
   AND con.ZONE = gen.ZONE
   AND fsk.ZONE = gen.ZONE
   and con.ZONE = imp_gen.ZONE (+)
   and con.mon= imp_gen.mon (+)
   and gen.mon=imp_con.mon (+)
   and gen.ZONE= imp_con.ZONE (+)
   ORDER BY con.zone
'''

Four_stor_sib = '''SELECT CASE WHEN con.zone = 1 THEN 'ДАЛЬНИЙ ВОСТОК' END ЗОНА,
            nvl(con,0) + nvl(gen,0) + 
nvl(fsk,0)+nvl(imp_gen,0)+nvl(imp_con,0) AS ОБЪЕМ, nvl(gen,0) AS 
ГЕНЕРАЦИЯ, nvl(con,0) AS 
ПОТРЕБЛЕНИЕ, nvl(fsk,0) AS ФСК, nvl(imp_gen,0) as ИМПОРТ , 
nvl(imp_con,0) as ЭКСПОРТ
  FROM (SELECT   TO_CHAR (a.target_date, 'MONTH') AS mon,
                 SUM (a.buy_volume) AS con, tr.is_unpriced_zone AS 
ZONE
            FROM frsdb_dev_sib.ncz_plan_con_volume a, 
frsdb_dev_sib.trader tr
           WHERE a.target_date BETWEEN tr.begin_date AND 
tr.end_date
             AND a.end_ver > 9999999
             AND a.target_date BETWEEN to_date(:d, 'dd.mm.yyyy') and last_day(to_date(:d, 'dd.mm.yyyy'))
             AND a.gtp_id = tr.real_trader_id
        GROUP BY TO_CHAR (a.target_date, 'MONTH'), 
tr.is_unpriced_zone) con,
       (SELECT   TO_CHAR (a.target_date, 'MONTH') AS mon,
                 SUM (a.sell_volume) AS gen, tr.is_unpriced_zone AS 
ZONE
            FROM frsdb_dev_sib.ncz_plan_gen_volume a, 
frsdb_dev_sib.trader tr
           WHERE a.target_date BETWEEN tr.begin_date AND 
tr.end_date
             AND a.end_ver > 9999999
             AND a.target_date BETWEEN to_date(:d, 'dd.mm.yyyy') and last_day(to_date(:d, 'dd.mm.yyyy'))
             AND a.gtp_id = tr.real_trader_id
        GROUP BY TO_CHAR (a.target_date, 'MONTH'), 
tr.is_unpriced_zone) gen,
       (SELECT   TO_CHAR (a.target_date, 'MONTH') AS mon, SUM 
(a.buy_volume)
                                                                       AS fsk,
                 tr.is_unpriced_zone AS ZONE
            FROM frsdb_dev_sib.ncz_plan_fsk_volume a, 
frsdb_dev_sib.trader tr
           WHERE a.target_date BETWEEN tr.begin_date AND 
tr.end_date
             AND a.end_ver > 9999999
             AND a.target_date BETWEEN to_date(:d, 'dd.mm.yyyy') and last_day(to_date(:d, 'dd.mm.yyyy'))
            and a.fsk_id=tr.real_trader_id
         GROUP BY TO_CHAR (a.target_date, 'MONTH'), 
tr.is_unpriced_zone) fsk,
          (SELECT   TO_CHAR (a.target_date, 'MONTH') AS mon, 
SUM (a.impex_volume)
                                                                       AS imp_gen,
                 tr.is_unpriced_zone AS ZONE
            FROM frsdb_dev_sib.ncz_plan_impex_volume a, 
frsdb_dev_sib.trader tr
           WHERE a.target_date BETWEEN tr.begin_date AND 
tr.end_date
             AND a.end_ver > 9999999
             AND a.target_date BETWEEN to_date(:d, 'dd.mm.yyyy') and last_day(to_date(:d, 'dd.mm.yyyy'))
             AND a.gtp_id = tr.real_trader_id
             and a.dir=2
         GROUP BY  TO_CHAR (a.target_date, 'MONTH'), 
tr.is_unpriced_zone) imp_gen,
          (SELECT   TO_CHAR (a.target_date, 'MONTH') AS mon, 
SUM (a.impex_volume)
                                                                       AS imp_con,
                 tr.is_unpriced_zone AS ZONE
            FROM frsdb_dev_sib.ncz_plan_impex_volume a, 
frsdb_dev_sib.trader tr
           WHERE a.target_date BETWEEN tr.begin_date AND 
tr.end_date
             AND a.end_ver > 9999999
             AND a.target_date BETWEEN to_date(:d, 'dd.mm.yyyy') and last_day(to_date(:d, 'dd.mm.yyyy'))
             AND a.gtp_id = tr.real_trader_id
             and a.dir=1
         GROUP BY  TO_CHAR (a.target_date, 'MONTH'), 
tr.is_unpriced_zone) imp_con
 WHERE con.mon = gen.mon
   AND fsk.mon = gen.mon
   AND con.ZONE = gen.ZONE
   AND fsk.ZONE = gen.ZONE
   and con.ZONE = imp_gen.ZONE (+)
   and con.mon= imp_gen.mon (+)
   and gen.mon=imp_con.mon (+)
   and gen.ZONE= imp_con.ZONE (+)
   ORDER BY con.zone
'''

# Формируем датафреймы и объединяем их
df_Four_stor_eur = pd.read_sql(Four_stor_eur, conn_eur, params={'d': date})
df_Four_stor_sib = pd.read_sql(Four_stor_sib, conn_sib, params={'d': date})

df_Four_stor = df_Four_stor_eur.append(df_Four_stor_sib)

# Записываем в xlsx
df_Four_stor.to_excel(path_FourSt_Zone, index=False)

# Создаем шаблоны стилей
border = Border(left=Side(border_style='thin', color='FF000000'),
                right=Side(border_style='thin', color='FF000000'),
                top=Side(border_style='thin', color='FF000000'),
                bottom=Side(border_style='thin', color='FF000000'))
align_head = Alignment(horizontal='center', vertical='center',
                       text_rotation=0, wrap_text=True,
                       shrink_to_fit=True, indent=0)
align_cell = Alignment(horizontal='right', vertical='center',
                       text_rotation=0, wrap_text=False,
                       shrink_to_fit=False, indent=0)

# Форматирование Excel
def exstyle(path):
    # Открываем рабочую страницу основного отчета
    wb = load_workbook(path)
    ws = wb.active
    # Устанавливаем ширину столбцов
    dim = {}
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            if cell.value:
                dim[cell.column] = max((dim.get(cell.column, 0), len(str(cell.value))))
    for col, value in dim.items():
        ws.column_dimensions[col].width = 8 + 0.85 * value
    # Задаем стили ячеек
    for row in ws.iter_rows():
        for cell in row:
            if cell in ws['1:1']:
                cell.border = border
                cell.alignment = align_head
            else:
                cell.border = border
                cell.alignment = align_cell
    wb.save(path)

exstyle(path_FourSt_Zone)


# Формирование письма
send_body = 'Добрый день!\n\nНаправляем отчеты по объемам договоров ' \
            'за ' + m[int(mon) - 1] + ' ' + ye + '.\n\n' \
            'С уважением, \nОтдел расчета объемов покупки и \nпродажи электрической энергии АО «АТС»'
msg = MIMEMultipart()
msg['From'] = cfg.send_from
msg['To'] = cfg.send_to_Four_stor_zone
msg['CC'] = cfg.send_cc
msg['Subject'] = '4-х сторонние договоры за ' + m[int(mon) - 1] + ' ' + ye
msg.attach(MIMEText(send_body, 'plain'))


def add_file(path):
    with open(path, "rb") as f:
        part = MIMEApplication(f.read())
    part.add_header('Content-Disposition', 'attachment', filename=os.path.basename(path))
    msg.attach(part)

add_file(path_FourSt_Zone)

# Отправка письма
s = smtplib.SMTP('smtp.rosenergo.com')
s.send_message(msg)
s.quit()

print('Письмо отправлено за: ', round(time() - time_start, 2), 'сек')