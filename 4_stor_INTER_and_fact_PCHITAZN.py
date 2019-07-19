import config as cfg
import cx_Oracle as ora
import pandas as pd
import os
from StyleFrame import StyleFrame, Styler, utils
from openpyxl.styles import Border, Side, Alignment
from openpyxl import load_workbook
from time import time

import smtplib
from email.mime.application import MIMEApplication
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText

time_start = time()

# Подключение к БД
conn_sib = ora.connect(cfg.user_db + '/' + cfg.pass_db + '@' + cfg.db_sib, encoding='utf-8')
conn_eur = ora.connect(cfg.user_db + '/' + cfg.pass_db + '@' + cfg.db_eur, encoding='utf-8')
# Загрузка даты отчета
ye = cfg.year
mon = cfg.mon
date = '01.' + mon + '.' + ye
print('Отчетный месяц: ', date)
# Список месяцев
m = ['январь', 'февраль', 'март', 'апрель', 'май', 'июнь',
     'июль', 'август', 'сентябрь', 'октябрь', 'ноябрь', 'декабрь']
# Корневая папка
path_0 = cfg.path
# Собираем путь к отчётам
path_1 = path_0 + 'Отчеты коллегам/' + ye + '/' + mon
path_inter = path_0 + 'Отчеты коллегам/' + ye + '/' + mon + '/' + '4-х сторонние договоры по экспорту-импорту' + ' ' + m[int(mon) - 1] + ' ' + ye + '.xlsx'
path_PCHITAZN = path_0 + 'Отчеты коллегам/' + ye + '/' + mon + '/' + 'Факт по PCHITAZN'+ ' ' + m[int(mon) - 1] + ' ' + ye + '.xlsx'

if not os.path.exists(path_1):
    os.makedirs(path_1)

# Загружаем исходные данные
Four_x_stor_INT_eur = ''' SELECT   'ОАО "ИНТЕР РАО ЕЭС"' AS УЧАСТНИК,
         CASE WHEN a.dir = 1 THEN 'PINTYANT' ELSE 'GINTYANT' END ГТП, 
            SUM (a.tg) AS ТГ, SUM (pl.iv) AS ИВ_ПЛЮС, SUM (mi.iv) AS ИВ_МИНУС,
            SUM (pl.is_volume) AS ИC_ПЛЮС, SUM (mi.is_volume) AS ИC_МИНУС
    FROM frsdb_dev.ncz_impex_volume a,
    
        (SELECT dir, target_date, hour,
                CASE WHEN iv > 0 THEN iv ELSE 0 END iv,
                CASE WHEN is_volume > 0 THEN is_volume ELSE 0 END is_volume
            FROM frsdb_dev.ncz_impex_volume
           WHERE end_ver > 9999999
             AND section_code IN ('WIMECHIN', 'WIMEYANT', 'WIMECHN1')
             AND is_daily = 0) pl,
             
        (SELECT dir, target_date, hour,
                CASE WHEN iv < 0 THEN - iv ELSE 0 END iv,
                CASE WHEN is_volume < 0 THEN - is_volume ELSE 0 END is_volume
            FROM frsdb_dev.ncz_impex_volume
           WHERE end_ver > 9999999
             AND section_code IN ('WIMECHIN', 'WIMEYANT', 'WIMECHN1')
             AND is_daily = 0) mi 
             
   WHERE a.target_date BETWEEN to_date(:d, 'dd.mm.yyyy') and last_day(to_date(:d, 'dd.mm.yyyy'))
     AND a.end_ver > 9999999
     AND pl.dir = a.dir
     AND mi.dir = a.dir
     AND pl.target_date = a.target_date
     AND mi.target_date = a.target_date
     AND pl.hour = a.hour
     AND mi.hour = a.hour
     AND a.section_code IN ('WIMECHIN', 'WIMEYANT', 'WIMECHN1')
     AND a.is_daily = 0
     GROUP BY a.dir'''

Four_x_stor_INT_sib = '''SELECT 'ОАО "ИНТЕР РАО ЕЭС"' AS УЧАСТНИК/*,
         CASE WHEN a.dir = 1 THEN 'PINTCHIN' ELSE 'GINTCHIN' END ГТП*/, t.trader_code as ГТП,
            SUM (a.tg) AS ТГ, SUM (pl.iv) AS ИВ_ПЛЮС, SUM (mi.iv) AS ИВ_МИНУС,
            SUM (pl.is_volume) AS ИC_ПЛЮС, SUM (mi.is_volume) AS ИC_МИНУС  
    FROM frsdb_dev_sib.ncz_impex_volume a,
    
        (SELECT dir, target_date, hour, gtp_id,
                CASE WHEN iv > 0 THEN iv ELSE 0 END iv,
                CASE WHEN is_volume > 0 THEN is_volume ELSE 0 END is_volume
            FROM frsdb_dev_sib.ncz_impex_volume
           WHERE end_ver > 99999999
           AND target_date BETWEEN to_date(:d, 'dd.mm.yyyy') and last_day(to_date(:d, 'dd.mm.yyyy'))
             AND section_code IN ('WIMECHIN', 'WIMEYANT', 'WIMECHN1')
             AND is_daily = 0) pl,
             
        (SELECT dir, target_date, hour, gtp_id,
                CASE WHEN iv < 0 THEN - iv ELSE 0 END iv,
                CASE WHEN is_volume < 0 THEN - is_volume ELSE 0 END is_volume
            FROM frsdb_dev_sib.ncz_impex_volume
           WHERE end_ver > 99999999
           AND target_date BETWEEN to_date(:d, 'dd.mm.yyyy') and last_day(to_date(:d, 'dd.mm.yyyy'))
             AND section_code IN ('WIMECHIN', 'WIMEYANT', 'WIMECHN1')
             AND is_daily = 0) mi,            
             frsdb_dev_sib.trader t            
   WHERE a.target_date BETWEEN to_date(:d, 'dd.mm.yyyy') and last_day(to_date(:d, 'dd.mm.yyyy'))
     AND a.end_ver > 99999999
     AND pl.dir = a.dir
     AND mi.dir = a.dir
     AND pl.target_date = a.target_date
     AND mi.target_date = a.target_date
     AND pl.hour = a.hour
     AND mi.hour = a.hour
     AND pl.gtp_id = mi.gtp_id
     AND a.gtp_id = pl.gtp_id
     AND a.section_code IN ('WIMECHIN', 'WIMEYANT', 'WIMECHN1')
     AND a.is_daily = 0
     AND a.target_date between t.begin_date AND t.end_date
     AND a.gtp_id = t.real_trader_id
     GROUP BY TRUNC (a.target_date, 'MONTH'), t.trader_code
'''

fact_PCHITAZN = '''select t.full_name as УЧАСТНИК , t.trader_code as ГТП , f.target_date as ДАТА,
                    f.hour as ЧАС, f.volume as ФАКТ
                     from frsdb_dev_sib.DEV_FOREM_FACT f, frsdb_dev_sib.trader t
                    where f.target_date between to_date(:d, 'dd.mm.yyyy') and last_day(to_date(:d, 'dd.mm.yyyy'))
                      and f.end_ver>9999999999999
                      and f.isdaily=0
                      and t.real_trader_id= f.oi_id
                      and t.trader_code = 'PCHITAZN'
                      and f.target_date between t.begin_date and t.end_date
                    order by t.trader_code, f.target_date, f.hour'''

df_Four_x_stor_INT_eur = pd.read_sql(Four_x_stor_INT_eur, conn_eur, params={'d': date})
df_Four_x_stor_INT_sib = pd.read_sql(Four_x_stor_INT_sib, conn_sib, params={'d': date})
df_fact_PCHITAZN = pd.read_sql(fact_PCHITAZN, conn_sib, params={'d': date})

df_fact_PCHITAZN['ДАТА'] = df_fact_PCHITAZN['ДАТА'].astype('str')

conn_eur.close()
conn_sib.close()

df_Four_x_stor_INT=df_Four_x_stor_INT_eur.append(df_Four_x_stor_INT_sib)

# Для PCHITAZN можно этот вариант форматирования массива и записи в xlsx
style = Styler(font_size=10,
               horizontal_alignment=utils.horizontal_alignments.right)

columns = df_fact_PCHITAZN.axes[1]
PCHITAZN = StyleFrame(df_fact_PCHITAZN)
excel_writer = StyleFrame.ExcelWriter(path_PCHITAZN)
for s in columns:
    PCHITAZN.set_column_width(s, 9 + str(s).__len__())

PCHITAZN.apply_column_style(cols_to_style=columns, styler_obj=style, style_header=True)
PCHITAZN.to_excel(excel_writer=excel_writer, index=False)
excel_writer.save()
excel_writer.close()

# Для склееного файла почему-то не катит предыдущий вариант форматирования массива данных, поэтому делаем так ->

# Экспортируем выгруженные данные в xlsx
df_Four_x_stor_INT.to_excel(path_inter, index=False)

# Создаем шаблоны стилей
border = Border(left=Side(border_style='thin', color='FF000000'),
                right=Side(border_style='thin', color='FF000000'),
                top=Side(border_style='thin', color='FF000000'),
                bottom=Side(border_style='thin', color='FF000000'))
align_head = Alignment(horizontal='center', vertical='center',
                       text_rotation=0, wrap_text=True,
                       shrink_to_fit=True, indent=0)
align_cell = Alignment(horizontal='right', vertical='center',
                       text_rotation=0, wrap_text=False,
                       shrink_to_fit=False, indent=0)

# Форматирование Excel
def exstyle(path):
    # Открываем рабочую страницу основного отчета
    wb = load_workbook(path)
    ws = wb.active
    # Устанавливаем ширину столбцов
    dim = {}
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            if cell.value:
                dim[cell.column] = max((dim.get(cell.column, 0), len(str(cell.value))))
    for col, value in dim.items():
        ws.column_dimensions[col].width = 8 + 0.85 * value
    # Задаем стили ячеек
    for row in ws.iter_rows():
        for cell in row:
            if cell in ws['1:1']:
                cell.border = border
                cell.alignment = align_head
            else:
                cell.border = border
                cell.alignment = align_cell
    wb.save(path)

exstyle(path_inter)

# Формирование письма
send_body = 'Добрый день!\n\nОтправляем список 4-х сторонних договоров ИНТЕР РАО и факт по ГТП PCHITAZN ' \
            'за ' + m[int(mon) - 1] + ' ' + ye + '.\n\n' \
            'С уважением, \nОтдел расчета объемов покупки и \nпродажи электрической энергии АО «АТС»'
msg = MIMEMultipart()
msg['From'] = cfg.send_from
msg['To'] = cfg.send_to_INTER_CHITA
msg['CC'] = cfg.send_cc
msg['Subject'] = '4-х сторонние договоры ИНТЕР РАО и факт по ГТП PCHITAZN ' + m[int(mon) - 1] + ' ' + ye
msg.attach(MIMEText(send_body, 'plain'))


def add_file(path):
    with open(path, "rb") as f:
        part = MIMEApplication(f.read())
    part.add_header('Content-Disposition', 'attachment', filename=os.path.basename(path))
    msg.attach(part)


add_file(path_PCHITAZN)
add_file(path_inter)
# Отправка письма
s = smtplib.SMTP('smtp.rosenergo.com')
s.send_message(msg)
s.quit()

print('Письмо отправлено за: ', round(time() - time_start, 2), 'сек')