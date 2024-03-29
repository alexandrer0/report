import config as cfg
import cx_Oracle as ora
import pandas as pd
import os
from openpyxl.styles import Border, Side, Alignment
from openpyxl import load_workbook
from time import time

time_start = time()

# Подключение к БД
conn_sib = ora.connect(cfg.user_db + '/' + cfg.pass_db + '@' + cfg.db_sib)
conn_eur = ora.connect(cfg.user_db + '/' + cfg.pass_db + '@' + cfg.db_eur)
# Загрузка даты отчета
ye = cfg.year
mon = cfg.mon
date = '01.' + mon + '.' + ye
print('Отчетный месяц: ', date)
# Список месяцев
m = ['январь', 'февраль', 'март', 'апрель', 'май', 'июнь',
     'июль', 'август', 'сентябрь', 'октябрь', 'ноябрь', 'декабрь']
# Корневая папка
path_0 = cfg.path
# Собираем путь к исходному реестру ДД
path_00 = path_0 + 'Данные для расчетов БР/' + ye + '/' + mon + \
          '/Реестр ДД в НЦЗ' + '/DDNCZ_reestr_01.' + mon + '.' + ye + '.xls'
# Собираем путь к папке с отчетом
path_1 = path_0 + 'Отчеты коллегам/' + ye + '/' + mon
# Собираем путь к папке с отчетом для мощности
path_2 = path_0 + 'Перетоки, отчет для мощности/' + ye + '/' + mon

# Проверяем наличие исходного реестра ДД
if not os.path.isfile(path_00):
    print('Прервано! Исходный реестр ДД не найден')
    raise SystemExit


# Создаем папки
def cfolder(path):
    if not os.path.exists(path):
        os.makedirs(path)


cfolder(path_1)
cfolder(path_2)
# Пути для отчетов
path_11 = path_1 + '/ddpost_' + mon + ye + '.xlsx'
path_1 += '/ДД в НЦЗ.xlsx'
path_3 = path_2 + '/VC_PC ' + m[int(mon) - 1] + ' ' + ye + '.xlsx'
path_4 = path_2 + '/Факт ИНТЕРРАО в НЦЗ за ' + m[int(mon) - 1] + ' ' + ye + '.xlsx'
path_2 += '/ДД ИНТЕРРАО в НЦЗ за ' + m[int(mon) - 1] + ' ' + ye + '.xlsx'

# Загружаем исходные данные
dd_0 = pd.read_excel(path_00)
query_dd_sib = '''select distinct v.dd_number, t.trader_code st_code, v.dd_fact, v.con_fact
from FRSDB_DEV_SIB.ncz_dd_volume v, frsdb_dev_sib.trader t
where v.target_date = to_date(:d, 'dd.mm.yyyy')
and v.end_ver=999999999999999 and t.real_trader_id=v.station_id
and v.target_date between t.begin_date and t.end_Date
order by 1,2'''
query_dd_eur = '''select distinct v.dd_number, t.trader_code st_code, v.dd_fact, v.con_fact
from FRSDB_DEV.ncz_dd_volume v, frsdb_dev.trader t
where v.target_date = to_date(:d, 'dd.mm.yyyy')
and v.end_ver=999999999999999 and t.real_trader_id=v.station_id
and v.target_date between t.begin_date and t.end_Date
order by 1,2'''
query_vc = '''select distinct start_ver,target_date, hour, gtp_id, section_code, impex_volume 
from frsdb_dev_sib.ncz_plan_impex_volume
where dir =1 and end_ver=999999999999999
and target_date between to_date(:d, 'dd.mm.yyyy') and last_day(to_date(:d, 'dd.mm.yyyy'))
order by section_code, target_date, hour'''
query_fact = '''select trunc(target_date, 'month') month, section_code, sum(fact) fact
from frsdb_dev_sib.ncz_impex_volume
where target_date between to_date(:d, 'dd.mm.yyyy') and last_day(to_date(:d, 'dd.mm.yyyy'))
and end_ver=999999999999999 and is_daily=0 and dir=1 
GROUP by trunc(target_date, 'month'), section_code'''

pf_sib = pd.read_sql(query_dd_sib, conn_sib, params={'d': date})
pf_eur = pd.read_sql(query_dd_eur, conn_eur, params={'d': date})
vc_pc = pd.read_sql(query_vc, conn_sib, params={'d': date})
sec_f = pd.read_sql(query_fact, conn_sib, params={'d': date})
conn_sib.close()
conn_eur.close()
# Создаем отчет
pf = pf_sib.append(pf_eur)
if pf.shape[0] == 0:
    print('Прервано! В БД отсутствуют данные по ДД')
    raise SystemExit
if vc_pc.shape[0] == 0 or sec_f.shape[0] == 0:
    print('Прервано! В БД отсутствуют данные по НЦЗ')
    raise SystemExit
pf.rename(columns={'DD_NUMBER': 'Номер ДД', 'ST_CODE': 'Код Станции Продавца'}, inplace=True)
dd_1 = dd_0.merge(pf, 'left', on=['Номер ДД', 'Код Станции Продавца'])
dd_1.rename(columns={'DD_FACT': 'ДД факт, кВтч', 'CON_FACT': 'Факт общий по ГТПП/ГТП экспорта, кВт'}, inplace=True)
dd_1.sort_values(['Номер ДД', 'Код Станции Продавца'], inplace=True)
# Экспортируем Excel
dd_1.to_excel(path_1, sheet_name='ДД в НЦЗ', index=False)
# Создаем отчет для ДФР
dd_1.drop('Факт общий по ГТПП/ГТП экспорта, кВт', axis=1, inplace=True)
dd_1.to_excel(path_11, sheet_name='dd_post', index=False)
# Создаем отчет по ИнтерРАО
colum = ['Объем ДД', 'Значение приоритета корректировки', 'Доля ГЭС', 'Доля ТЭС', 'Номер пакета ДД']
dd_1.drop(colum, axis=1, inplace=True)
dd_1.drop(dd_1[dd_1['Код ГТПП Покупателя'] != 'PINTCHIN PINTCHN1'].index, axis=0, inplace=True)
vc_pc['TARGET_DATE'] = vc_pc['TARGET_DATE'].astype(str)
sec_f['MONTH'] = sec_f['MONTH'].astype(str)
# Экспортируем Excel
dd_1.to_excel(path_2, sheet_name='ДД ИНТЕРРАО в НЦЗ', index=False)
vc_pc.to_excel(path_3, sheet_name='VC_PC', index=False)
sec_f.to_excel(path_4, sheet_name='Факт ИНТЕРРАО в НЦЗ', index=False)
# Создаем шаблоны стилей
border = Border(left=Side(border_style='thin', color='FF000000'),
                right=Side(border_style='thin', color='FF000000'),
                top=Side(border_style='thin', color='FF000000'),
                bottom=Side(border_style='thin', color='FF000000'))
align_head = Alignment(horizontal='center', vertical='center',
                       text_rotation=0, wrap_text=True,
                       shrink_to_fit=True, indent=0)
align_cell = Alignment(horizontal='right', vertical='center',
                       text_rotation=0, wrap_text=False,
                       shrink_to_fit=False, indent=0)


# Форматирование Excel
def exstyle(path):
    # Открываем рабочую страницу основного отчета
    wb = load_workbook(path)
    ws = wb.active
    # Устанавливаем ширину столбцов
    dim = {}
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            if cell.value:
                dim[cell.column] = max((dim.get(cell.column, 0), len(str(cell.value))))
    for col, value in dim.items():
        ws.column_dimensions[col].width = 11 + 0.85 * value
    # Задаем стили ячеек
    for row in ws.iter_rows():
        for cell in row:
            if cell in ws['1:1']:
                cell.border = border
                cell.alignment = align_head
            else:
                cell.border = border
                cell.alignment = align_cell
    wb.save(path)


exstyle(path_1)
exstyle(path_11)
exstyle(path_2)
exstyle(path_3)
exstyle(path_4)

print('Отчеты созданы успешно за: ', round(time() - time_start, 2), 'сек')
